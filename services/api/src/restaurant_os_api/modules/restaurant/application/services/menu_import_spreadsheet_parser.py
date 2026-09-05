"""Deterministic CSV/XLSX parsing for the menu-import feature.

Unlike photos/PDFs, a spreadsheet is already structured -- there is no
reason to pay for a vision call on it. This produces the exact same
``ExtractedMenuRowDTO`` shape the vision extractor does (before price
normalization, which both paths defer to the use case), so the review
grid downstream doesn't care which pipeline a row came from.

Column headers are matched case-insensitively against a small set of
known synonyms rather than requiring an exact schema -- a menu export
is never going to consistently use one column-naming convention. A
column with no synonym match just isn't populated for that field
(category defaults to "Uncategorized", dietary/portion/unit default to
unknown/empty) rather than failing the whole file, except name and
price: without those two, there's nothing to review, so that raises
``MenuImportExtractionFailedError`` instead of silently returning zero
usable rows.
"""

from __future__ import annotations

import csv
import io

from openpyxl import load_workbook

from restaurant_os_api.modules.restaurant.application.dto import (
    ExtractedMenuRowDTO,
    MenuImportConfidence,
)
from restaurant_os_api.modules.restaurant.domain.exceptions import (
    MenuImportExtractionFailedError,
)

_CATEGORY_HEADERS = {"category", "section", "group", "heading"}
_NAME_HEADERS = {"item", "name", "dish", "item name", "dish name", "item name "}
_PRICE_HEADERS = {"price", "rate", "amount", "mrp", "cost"}
_PORTION_HEADERS = {"portion", "size", "portion size"}
_UNIT_HEADERS = {"unit", "pricing unit", "per"}
_DIETARY_HEADERS = {"veg", "dietary", "type", "dietary type", "veg/non-veg"}

_UNCATEGORIZED = "Uncategorized"


def _match_header(headers: list[str], synonyms: set[str]) -> str | None:
    for header in headers:
        if header.strip().lower() in synonyms:
            return header
    return None


def _rows_from_dicts(rows: list[dict[str, object]]) -> list[ExtractedMenuRowDTO]:
    if not rows:
        raise MenuImportExtractionFailedError("the file has no data rows")

    headers = list(rows[0].keys())
    name_col = _match_header(headers, _NAME_HEADERS)
    price_col = _match_header(headers, _PRICE_HEADERS)
    if name_col is None or price_col is None:
        raise MenuImportExtractionFailedError(
            "couldn't find item name / price columns -- expected headers like "
            "'Item'/'Name' and 'Price'/'Rate'/'Amount'"
        )
    category_col = _match_header(headers, _CATEGORY_HEADERS)
    portion_col = _match_header(headers, _PORTION_HEADERS)
    unit_col = _match_header(headers, _UNIT_HEADERS)
    dietary_col = _match_header(headers, _DIETARY_HEADERS)

    result: list[ExtractedMenuRowDTO] = []
    for index, row in enumerate(rows):
        name = str(row.get(name_col) or "").strip()
        raw_price = str(row.get(price_col) or "").strip()
        if not name and not raw_price:
            continue  # a genuinely blank row (trailing blank lines etc.)

        category = str(row.get(category_col) or "").strip() if category_col else ""
        confidence = MenuImportConfidence.HIGH
        note = None
        if not name or not raw_price:
            confidence = MenuImportConfidence.LOW
            note = "Missing name or price in this row."
        if not category:
            category = _UNCATEGORIZED

        result.append(
            ExtractedMenuRowDTO(
                category=category,
                name=name,
                raw_price=raw_price,
                price_amount=None,  # normalized later, uniformly, by the use case
                confidence=confidence,
                source_image_index=None,
                dietary_type=(str(row.get(dietary_col) or "").strip() or None)
                if dietary_col
                else None,
                portion_label=(str(row.get(portion_col) or "").strip() or None)
                if portion_col
                else None,
                pricing_unit=(str(row.get(unit_col) or "").strip() or None) if unit_col else None,
                note=note,
            )
        )

    if not result:
        raise MenuImportExtractionFailedError("no usable rows found in the file")
    return result


def parse_csv(data: bytes) -> list[ExtractedMenuRowDTO]:
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return _rows_from_dicts([dict(row) for row in reader])


def parse_xlsx(data: bytes) -> list[ExtractedMenuRowDTO]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise MenuImportExtractionFailedError("the file has no data rows") from exc

    headers = [str(cell) if cell is not None else "" for cell in header_row]
    dict_rows: list[dict[str, object]] = []
    for row in rows_iter:
        if all(cell is None for cell in row):
            continue
        dict_rows.append(dict(zip(headers, row, strict=False)))
    return _rows_from_dicts(dict_rows)
